import streamlit as st

import pandas as pd
import numpy as np

from io import BytesIO

from typing import Tuple, List, Union, Optional
from pathlib import Path

import zipfile
import tempfile
import os

from openpyxl import load_workbook, Workbook

import copy

from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

class ExcelProcessor:
    """
    A class to process and manage Excel data related to BCA (Budget Control Activities), 
    assets, and purchase order details.

    Attributes:
        bca_path (str): File path to the BCA Excel file.
        assets_path (str): File path to the assets Excel file.
        po_details_path (str): File path to the purchase order details Excel file.
        df_bca (pd.DataFrame): DataFrame to store the BCA data.
        df_assets (pd.DataFrame): DataFrame to store the assets data.
        df_po (pd.DataFrame): DataFrame to store the purchase order data.
        df_sorted (pd.DataFrame): DataFrame to store processed or sorted data.
        balances (bool): Flag indicating if balances are being considered.
        required_columns (List[str]): List of column names expected in the data.
        sheet_names (List[str]): List of sheet names to load from Excel files.
        file_paths (List[str]): List of file paths for processing.
        CFB_flag (int): Internal flag for processing status or condition.
    """

    def __init__(
        self,
        bca_path: str = '',
        assets_path: str = '',
        po_details_path: str = ''
    ):
        self.bca_path = bca_path
        self.assets_path = assets_path
        self.po_details_path = po_details_path
        self.df_bca = pd.DataFrame()
        self.df_assets = pd.DataFrame()
        self.df_po = pd.DataFrame()
        self.df_sorted = pd.DataFrame()
        self.balances = False
        self.required_columns = ['Transaction Number', 'Balance Type', 'Transaction Amount', 
                    'Liquidation Transaction Number', 'Cluster', 
                    'Commitment Nr', 'Obligation Nr', 'Expenditure Nr']
        
        self.sheet_names = ['BCA']
        self.file_paths = []

        self.CFB_flag = 0


    def extract_data_from_excel(
        self,
        file_path: Union[str, Path],
        file_type: str = 'bca',
        verbose: bool = False
    ) -> Optional[pd.DataFrame]:
        """
        Extracts data from an Excel file based on the specified file type.

        This function attempts to locate a target header value in the first column of the first sheet.
        Once found, it reads the table from that row onward into a DataFrame.
        Additional preprocessing is done for 'bca' file types, including adding empty columns
        and converting data types.

        Parameters
        ----------
        file_path : str or Path
            Path to the Excel file to read.
        file_type : str, default 'bca'
            The type of file to extract. Supported: 'bca', 'PO'.
            This affects the target header string and some processing steps.
        verbose : bool, default False
            If True, prints additional debugging information about the loaded data.

        Returns
        -------
        Optional[pandas.DataFrame]
            Returns a DataFrame if data is successfully extracted; otherwise None.

        Notes
        -----
        - Uses openpyxl to read the Excel file.
        - Warns via Streamlit if no data is found.
        """
        print('Extracting')
        workbook = load_workbook(filename=file_path)

        found_target = False        

        # Set target header value based on file type
        target_value = ''
        if file_type == 'bca':
            target_value = 'Budget Account'
        elif file_type == 'PO':
            target_value = 'Procurement Business Unit'

        print('Loop')
        loop_count = 0
        while not found_target and loop_count < 5:
            loop_count += 1
            sheet = workbook.worksheets[0]
            num_rows = sheet.max_row
            target_row_idx = -1
            target_col_idx = 1  # First column

            # Search for target_value in first column
            for row_idx in range(1, num_rows + 1): 
                cell_value = sheet.cell(row=row_idx, column=target_col_idx).value
                if cell_value == target_value:
                    found_target = True
                    target_row_idx = row_idx
                    break

            if found_target:
                # Extract header row values
                headers = [cell.value for cell in sheet[target_row_idx]]

                # Extract data rows below header
                start_row_idx = target_row_idx + 1
                data = []
                for row in sheet.iter_rows(min_row=start_row_idx, max_row=num_rows, values_only=True):
                    data.append(row)

                df = pd.DataFrame(data, columns=headers)

            else:
                # Adjust target_value if not found (for PO file_type)
                if file_type == 'PO':
                    target_value = 'Procurement Business Unit Name'
        
        print('Loop Done')

        # If data found, check if DataFrame is empty
        if found_target:
            all_none = df.isnull().all().all()
        else:
            all_none = True
            st.warning(f"**_Warning:_** No data found in {Path(file_path).name}.")

        if all_none:
            return None
        else:
            if file_type == 'bca':
                # Add empty columns for later processing
                df['Cluster'] = ''
                df['Commitment Nr'] = ''
                df['Obligation Nr'] = ''
                df['Expenditure Nr'] = ''
                df['Requester Name'] = ''
                df['Supplier Name'] = ''
                df['Item Description'] = ''
                df['Project Code'] = ''
                df['Item Category Description'] = ''

                # Convert Transaction Amount to float (handle commas)
                try:
                    df['Transaction Amount'] = df['Transaction Amount'].astype(str).str.replace(',', '')
                    df['Transaction Amount'] = df['Transaction Amount'].astype(float)
                    df['Transaction Number'] = df['Transaction Number'].astype(str)
                except Exception:
                    print('Error converting Transaction Amount to float')

                # Split dataframe by Balance Type for summary statistics
                df_obligations = df[df['Balance Type'] == 'Obligation']
                df_commitments = df[df['Balance Type'] == 'Commitment']
                df_budget = df[df['Balance Type'] == 'Budget']
                df_expenditures = df[df['Balance Type'] == 'Expenditure']

                # Check for missing required columns
                self.missing_columns = [col for col in self.required_columns if col not in df.columns]

                if not self.missing_columns:
                    # All required columns present
                    pass
                else:
                    st.markdown("The following columns are missing:", self.missing_columns)

                if verbose:
                    print('Before operations:\n')
                    print(f'Obligations: {df_obligations.shape}')
                    print(f'Commitments: {df_commitments.shape}')
                    print(f'Budget: {df_budget.shape}')
                    print(f'Expenditures: {df_expenditures.shape}')
                    total = (df_obligations.shape[0] + df_commitments.shape[0] + 
                            df_budget.shape[0] + df_expenditures.shape[0])
                    print(f'Total rows: {total}\n')

                    print('Sum of transaction amounts by category:')
                    print(f"Obligation Sum: {df_obligations['Transaction Amount'].sum()}")
                    print(f"Commitment Sum: {df_commitments['Transaction Amount'].sum()}")
                    print(f"Budget Sum: {df_budget['Transaction Amount'].sum()}")
                    print(f"Expenditure Sum: {df_expenditures['Transaction Amount'].sum()}")

                return df
            else:
                # For other file types, simply return the DataFrame
                return df


    def add_descriptions(self, df: pd.DataFrame, numbers: list[int]) -> pd.DataFrame:
        """
        Adds description columns ('Commitment Nr', 'Obligation Nr', 'Expenditure Nr') to the DataFrame
        based on transaction and liquidation transaction relationships.

        For each transaction number in `numbers`, the method:
        - Finds related liquidation transactions (obligations).
        - Finds related expenditures tied to those obligations.
        - Assigns commitment, obligation, and expenditure numbers accordingly in the DataFrame.

        Args:
            df (pd.DataFrame): Input DataFrame with columns such as 
                            'Transaction Number', 'Liquidation Transaction Number', 'Balance Type'.
            numbers (list[int]): List of transaction numbers to process.

        Returns:
            pd.DataFrame: The input DataFrame updated with the new description columns.
        """
        for i in range(len(numbers)):
            comm_nr = [numbers[i]]
            obl_nr = df.loc[
                (df['Liquidation Transaction Number'] == comm_nr[0]),
                'Transaction Number'
            ].unique()

            if len(obl_nr) == 0:
                df.loc[
                    (df['Transaction Number'] == comm_nr[0]),
                    'Commitment Nr'
                ] = comm_nr[0]
            else:
                for obl_number in obl_nr:
                    exp_nr = df.loc[
                        (df['Liquidation Transaction Number'] == obl_number) & 
                        (df['Balance Type'] != 'Commitment'),
                        'Transaction Number'
                    ].unique()
                    
                    if len(exp_nr) == 0:
                        df.loc[
                            (df['Transaction Number'] == obl_number),
                            'Commitment Nr'
                        ] = comm_nr[0]

                        df.loc[
                            (df['Transaction Number'] == obl_number),
                            'Obligation Nr'
                        ] = obl_number
                    else:
                        for exp_number in exp_nr:
                            transaction_number = np.concatenate((comm_nr, [obl_number], [exp_number]))

                            df.loc[
                                (df['Transaction Number'].isin(transaction_number)) & 
                                (df['Balance Type'] == 'Expenditure'),
                                'Commitment Nr'
                            ] = comm_nr[0]

                            df.loc[
                                (df['Transaction Number'].isin(transaction_number)) & 
                                (df['Balance Type'] == 'Expenditure'),
                                'Obligation Nr'
                            ] = obl_number

                            df.loc[
                                (df['Transaction Number'].isin(transaction_number)) & 
                                (df['Balance Type'] == 'Expenditure'),
                                'Expenditure Nr'
                            ] = exp_number

        return df


    def group_and_sort(self, df: pd.DataFrame, numbers: list[int], threshold: float = 10, offset: int = 0) -> pd.DataFrame:
        """
        Groups transactions into clusters and computes aggregated amounts for commitments and obligations.
        Also assigns project codes based on amount thresholds.

        Steps:
        - Assigns a cluster number to transactions related to each number in `numbers`.
        - Aggregates the 'Transaction Amount' by cluster and 'Balance Type' (Commitment or Obligation) into a new 'Amount' column.
        - Marks transactions with amounts between -threshold and threshold as 'Ignore' in the 'Project Code' column.

        Args:
            df (pd.DataFrame): DataFrame containing transactions with columns including 'Transaction Number',
                            'Liquidation Transaction Number', 'Balance Type', and 'Transaction Amount'.
            numbers (list[int]): List of transaction numbers to cluster.
            threshold (float, optional): Threshold to classify small amounts as 'Ignore'. Defaults to 10.
            offset (int, optional): Offset for cluster numbering to avoid collisions. Defaults to 0.

        Returns:
            pd.DataFrame: The DataFrame updated with 'Cluster', 'Amount', and 'Project Code' columns.
        """   

        # Add cluster numbers
        for i in range(len(numbers)):
            trans_nrs = [numbers[i]]
            trans_nrs = df.loc[df['Liquidation Transaction Number'].isin(trans_nrs) | 
                df['Transaction Number'].isin(trans_nrs), 'Transaction Number'].unique()
            
            trans_nrs = df.loc[df['Liquidation Transaction Number'].isin(trans_nrs) | 
                df['Transaction Number'].isin(trans_nrs),'Transaction Number'].unique()

            df.loc[df['Transaction Number'].isin(trans_nrs), 'Cluster'] = offset + i

        # Create the transaction amounts based on cluster
        for i in range(len(numbers)):
            df.loc[
                (df['Cluster'] == i) & 
                (df['Balance Type'] == 'Commitment'), 
                'Amount'
            ] = df.loc[
                (df['Cluster'] == i) & 
                (df['Balance Type'] == 'Commitment'), 
                'Transaction Amount'
            ].sum()

            df.loc[
                (df['Cluster'] == i) & 
                (df['Balance Type'] == 'Obligation'), 
                'Amount'
            ] = df.loc[
                (df['Cluster'] == i) & 
                (df['Balance Type'] == 'Obligation'), 
                'Transaction Amount'
            ].sum()

        # Assign project codes based on amount vs threshold
        df.loc[
            (df['Amount'] < threshold) & 
            (df['Amount'] > -threshold), 
            'Project Code'
        ] = 'Ignore'

        return df


    def concatenate_rows_by_po_number(self, df: pd.DataFrame, group_by: str = 'Purchase Order Number') -> pd.DataFrame:
        """
        Groups the DataFrame by the specified column and concatenates the string representations
        of each group’s values, separated by ' | '.

        Args:
            df (pd.DataFrame): Input DataFrame.
            group_by (str, optional): Column name to group by. Defaults to 'Purchase Order Number'.

        Returns:
            pd.DataFrame: Aggregated DataFrame with concatenated string values per group.
        """
        return df.groupby(group_by).agg(lambda x: ' | '.join(x.astype(str))).reset_index()  


    def create_file_name(self, filename: str) -> str:
        """
        Formats a filename by keeping the first two parts separated by " - " and appending " - Output.xlsx".

        Args:
            filename (str): Original filename string.

        Returns:
            str: Formatted filename string.
        """       
        formatted_filename = " - ".join(filename.split(" - ")[:2]) + " - Output.xlsx"

        return formatted_filename
    

    def create_output_file(self, df: pd.DataFrame, file_paths: List[str]) -> BytesIO:
        """
        Combines original Excel files and appends processed data and balances into a new Excel workbook.
        
        This function:
        - Copies sheets from input files in `file_paths` into a new workbook
        - Adds a 'Processed' sheet with the given DataFrame `df`
        - Adds a 'Balances' sheet generated from `self.get_balances()`
        - Applies formatting, adjusts column widths, sets auto-filters and frozen panes
        - Updates `progress_placeholder` at multiple checkpoints

        Args:
            df (pd.DataFrame): The processed transaction data to include in the 'Processed' sheet.
            file_paths (List[str]): List of file paths to the original Excel input files.

        Returns:
            BytesIO: A stream of the created Excel file, ready to be written to disk or returned via web interface.
        """
        new_workbook = Workbook()
        new_workbook.remove(new_workbook.active)  # Remove default empty sheet

        # Copy each original sheet from the input files
        for i, path in enumerate(file_paths):
            workbook = load_workbook(path)
            sheet = workbook.active
            new_sheet = new_workbook.create_sheet(title=self.sheet_names[i])

            for row in sheet:
                for cell in row:
                    new_cell = new_sheet[cell.coordinate]
                    new_cell.value = cell.value
                    new_cell.font = copy.copy(cell.font)
                    new_cell.border = copy.copy(cell.border)
                    new_cell.fill = copy.copy(cell.fill)
                    new_cell.number_format = copy.copy(cell.number_format)
                    new_cell.protection = copy.copy(cell.protection)
                    new_cell.alignment = copy.copy(cell.alignment)

            for merged_range in sheet.merged_cells.ranges:
                new_sheet.merge_cells(str(merged_range))

        progress_placeholder.markdown("Current processing: 70% complete...")

        # Add 'Processed' sheet with formatted df
        new_sheet = new_workbook.create_sheet(title='Processed')

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                new_cell = new_sheet.cell(row=r_idx, column=c_idx, value=value)

                # Bold header row
                if r_idx == 1:
                    new_cell.font = Font(bold=True)

                # Format currency column (24 = 'Transaction Amount')
                if c_idx == 24:
                    new_cell.number_format = 'R#,##0.00'

        # Auto-adjust column widths based on content
        for col in new_sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = max_length + 2
            new_sheet.column_dimensions[column].width = adjusted_width

        # Override fixed widths for first 16 columns
        for col_idx in range(1, 17):
            col_letter = get_column_letter(col_idx)
            new_sheet.column_dimensions[col_letter].width = 12.75

        # Freeze header row and enable filter
        new_sheet.auto_filter.ref = new_sheet.dimensions
        new_sheet.freeze_panes = 'A2'

        progress_placeholder.markdown("Current processing: 80% complete...")

        # Add 'Balances' sheet with results from get_balances()
        df_balances = self.get_balances()
        new_sheet = new_workbook.create_sheet(title='Balances')

        for r_idx, row in enumerate(dataframe_to_rows(df_balances, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                new_cell = new_sheet.cell(row=r_idx, column=c_idx, value=value)

                # Bold headers
                if r_idx == 1:
                    new_cell.font = Font(bold=True)

                # Format currency columns (B, C, D => 2, 3, 4)
                if c_idx in [2, 3, 4]:
                    new_cell.number_format = 'R#,##0.00'

        # Set fixed column widths for Balances
        new_sheet.column_dimensions['A'].width = 80
        new_sheet.column_dimensions['B'].width = 25
        new_sheet.column_dimensions['C'].width = 25
        new_sheet.column_dimensions['D'].width = 25

        progress_placeholder.markdown("Current processing: 90% complete...")

        # Save workbook to an in-memory BytesIO stream
        excel_stream = BytesIO()
        new_workbook.save(excel_stream)
        excel_stream.seek(0)

        return excel_stream


    def get_balances(self) -> pd.DataFrame:
        """
        Constructs a DataFrame representing financial balance formulas to be used in an Excel sheet.

        The DataFrame contains descriptions and Excel formula strings for three columns:
        'BCA', 'BCA Assets', and 'Total'. The formulas dynamically reference other sheets
        such as 'Processed', 'BCA', and 'BCA Assets' based on the object's state flags.

        The logic adapts based on:
        - `self.CFB_flag` (int) which affects the formulas used in the 'BCA' and 'BCA Assets' columns.
        - `self.balances` (bool) which determines if formulas for 'BCA Assets' are actual formulas or zeroes.

        Returns:
            pd.DataFrame: A DataFrame with 7 rows and columns: 
                        ['Description', 'BCA', 'BCA Assets', 'Total'], 
                        where 'BCA', 'BCA Assets', and 'Total' columns contain Excel formulas as strings.
        """
        data = {
                'Description': [
                'Period opening balances (from reports):',
                'Period closing balances (available budget):',
                'Commitments during period (from reports):',
                'Obligations during period (from reports):',
                'Expenses during period (calculated):',
                'Total consumption during period (from reports - calculated total income):',
                'Total income during period (calculated):'
            ],
            'BCA': [None] * 7 ,
            'BCA Assets': [None] * 7,
            'Total': [None] * 7}
        df = pd.DataFrame(data)

        if self.CFB_flag >= 1:
            formulas_bca = [
                "=Processed!X2",
                "=BCA!K4",
                "=BCA!AA23",
                "=BCA!AD23",
                "=-(B3-B2-B8-B4-B5)",
                "=BCA!AI23",
                '=-SUMIF(Processed!W:W, "Income", Processed!X:X)'
            ]
        else:
            formulas_bca = [
                "=BCA!D4",
                "=BCA!K4",
                "=BCA!AA23",
                "=BCA!AD23",
                "=-(B3-B2-B8-B4-B5)",
                "=BCA!AI23",
                '=-SUMIF(Processed!W:W, "Income", Processed!X:X)'
            ]

        if self.balances:
            if self.CFB_flag == 2:
                formulas_assets = [
                    "=Processed!X3",
                    "=IF(ISNUMBER('BCA Assets'!K4),'BCA Assets'!K4,0)",
                    "=IF(ISNUMBER('BCA Assets'!AA23),'BCA Assets'!AA23,0)",
                    "=IF(ISNUMBER('BCA Assets'!AC23),'BCA Assets'!AC23,0)",
                    "=-(C3-C2-C8-C4-C5)",
                    "=IF(ISNUMBER('BCA Assets'!G4),'BCA Assets'!AK42+'BCA Assets'!AK41+'BCA Assets'!AK34+'BCA Assets'!AK35,0)",
                    "=0"
                ]
            else:
                formulas_assets = [
                    "=IF(ISNUMBER('BCA Assets'!D4),'BCA Assets'!D4,0)",
                    "=IF(ISNUMBER('BCA Assets'!K4),'BCA Assets'!K4,0)",
                    "=IF(ISNUMBER('BCA Assets'!AA23),'BCA Assets'!AA23,0)",
                    "=IF(ISNUMBER('BCA Assets'!AC23),'BCA Assets'!AC23,0)",
                    "=-(C3-C2-C8-C4-C5)",
                    "=IF(ISNUMBER('BCA Assets'!G4),'BCA Assets'!AK42+'BCA Assets'!AK41+'BCA Assets'!AK34+'BCA Assets'!AK35,0)",
                    "=0"
                ]
        else:
            print('Entered')
            formulas_assets = [
                "=0",
                "=0",
                "=0",
                "=0",
                "=0",
                "=0",
                "=C3-(C2-C7)"
            ]

        formulas_total = [
            "=B2+C2",
            "=B3+C3",
            "=B4+C4",
            "=B5+C5",
            "=B6+C6",
            "=B7+C7",
            "=B8+C8"
        ]

        # Update the DataFrame with the formulas
        for i, formula in enumerate(formulas_bca):
            df.at[i, 'BCA'] = formula
    
        for i, formula in enumerate(formulas_assets):
            df.at[i, 'BCA Assets'] = formula

        for i, formula in enumerate(formulas_total):
            df.at[i, 'Total'] = formula

        return df


    def extract_and_check_data(self) -> Tuple[bool, bool, bool]:
        """
        Extracts data from the provided BCA, Assets, and PO Details Excel file paths.
        Performs checks to determine whether the files contain valid data and provides
        warnings if any required data is missing.

        Returns:
            Tuple[bool, bool, bool]: A tuple of three booleans indicating whether
            data was successfully extracted from the BCA file, Assets file, and PO Details file respectively.
        """
        # Extract data from the BCA file
        if self.bca_path is not None:
            self.df_bca = self.extract_data_from_excel(self.bca_path, verbose=False)
        else:
            self.df_bca = None

        # Extract data from the Assets file if provided
        if self.assets_path is not None:
            self.df_assets = self.extract_data_from_excel(self.assets_path, verbose=False)
        else:
            self.df_assets = None

        # Extract data from the PO Details file if provided
        if self.po_details_path is not None:
            self.df_po = self.extract_data_from_excel(self.po_details_path, file_type='PO', verbose=False)
        else:
            self.df_po = None

        # Check if the extracted data is None
        bca_exists = self.df_bca is not None
        assets_exists = self.df_assets is not None
        po_exists = self.df_po is not None

        # Provide feedback based on the existence of the data
        if not bca_exists:
            st.warning('BCA file is empty. Please ensure that the SUNFIN Sheet has transactions.')
        if (not assets_exists) & (self.assets_path is not None):
            st.warning('BCA Assets file is empty.')
        if (self.po_details_path is not None) & (not po_exists):
            st.warning('PO Details file is empty.')

        # Return a tuple of True/False for each file
        return bca_exists, assets_exists, po_exists


    # 10% Completion
    def process_bca(self) -> None:
        """
        Processes the BCA (Budget Control Account) Excel data to clean and prepare it for analysis.

        Steps include:
        - Updating the file path tracking list.
        - Identifying and relabeling opening balance transactions from known patterns.
        - Converting specific 'Expenditure' transactions to 'Income' based on account codes.
        - Reclassifying adjustment receipts and budget increases as 'Income'.
        - Negating the transaction amounts for income entries.
        - Removing budget entries and cleaning up unnamed columns.

        This method updates `self.df_bca` in-place.
        """

        # Add BCA sheet file path to the lists
        self.file_paths.append(self.bca_path)

        # Find opening balances in the BCA data
        self.df_bca.loc[
            (self.df_bca['Balance Type'] == 'Budget') &
            self.df_bca['Transaction Number'].str.contains("Carry forward balance", na=False), 'Balance Type'] = 'Opening Balance'
        
        if 'Opening Balance' not in self.df_bca['Balance Type'].unique():
            self.df_bca.loc[
                (self.df_bca['Balance Type'] == 'Budget') &
                self.df_bca['Transaction Number'].str.contains("Initial Budget_CF", na=False), 'Balance Type'] = 'Opening Balance'
        
        if 'Opening Balance' not in self.df_bca['Balance Type'].unique():
            self.df_bca.loc[
                (self.df_bca['Balance Type'] == 'Budget') &
                self.df_bca['Transaction Number'].str.contains("Temporary budget_TERA", na=False), 'Balance Type'] = 'Opening Balance'

        # Convert expenditure transactions to income where applicable
        mask = (self.df_bca['Balance Type'] == 'Expenditure') & (self.df_bca['Transaction Account'].str.contains("5227", na=False))
        self.df_bca.loc[mask, 'Transaction Amount'] = self.df_bca.loc[mask, 'Transaction Amount'] 
        self.df_bca.loc[mask, 'Balance Type'] = 'Income'
        
        self.df_bca.loc[(self.df_bca['Transaction Number'].str.contains("Adjustments Receipts", na=False)) | 
                        (self.df_bca['Transaction Number'].str.contains("Adjustment: budget increase", na=False)), 'Balance Type'] = 'Income'
        
        # Convert income transactions to negative amounts
        self.df_bca.loc[self.df_bca['Balance Type'] == 'Income', 'Transaction Amount'] = self.df_bca.loc[self.df_bca['Balance Type'] == 'Income', 'Transaction Amount'] * -1

        # Filter out budget transactions and drop columns with None values
        self.df_bca = self.df_bca.loc[(self.df_bca['Balance Type'] != 'Budget')]
        self.df_bca = self.df_bca.drop(columns=[col for col in self.df_bca.columns if col is None])


    # 20% Completion
    def process_assets(self) -> None:
        """
        Processes the BCA Assets Excel data by identifying and adjusting financial classifications
        and merging it into the main BCA dataset.

        Steps include:
        - Tracking the sheet and file paths.
        - Identifying and relabeling opening balance transactions in the assets.
        - Reclassifying specific 'Expenditure' transactions to 'Income'.
        - Reclassifying adjustment receipts and budget increases as 'Income'.
        - Making income transactions negative in amount.
        - Dropping budget entries and cleaning unnamed columns.
        - Concatenating processed asset data into the main `df_bca` DataFrame.

        This method updates `self.df_assets` and merges it into `self.df_bca`.
        """

        # Add BCA Assets sheet file path to the lists
        self.sheet_names.append('BCA Assets')
        self.file_paths.append(self.assets_path)

        # Find opening balances in the BCA Assets data
        self.df_assets.loc[
            (self.df_assets['Balance Type'] == 'Budget') &
            self.df_assets['Transaction Number'].str.contains("Carry forward balance", na=False), 'Balance Type'] = 'Opening Balance Assets'
        
        if 'Opening Balance Assets' not in self.df_assets['Balance Type'].unique():
            self.df_assets.loc[
                (self.df_assets['Balance Type'] == 'Budget') &
                self.df_assets['Transaction Number'].str.contains("Initial Budget_CF", na=False), 'Balance Type'] = 'Opening Balance Assets'
        
        if 'Opening Balance Assets' not in self.df_assets['Balance Type'].unique():
            self.df_assets.loc[
                (self.df_assets['Balance Type'] == 'Budget') &
                self.df_assets['Transaction Number'].str.contains("Temporary budget_TERA", na=False), 'Balance Type'] = 'Opening Balance Assets'
        
        # Convert expenditure transactions to income where applicable
        mask = (self.df_assets['Balance Type'] == 'Expenditure') & (self.df_assets['Transaction Account'].str.contains("5227", na=False))
        self.df_assets.loc[mask, 'Transaction Amount'] = self.df_assets.loc[mask, 'Transaction Amount'] 
        self.df_assets.loc[mask, 'Balance Type'] = 'Income'
        
        self.df_assets.loc[(self.df_assets['Transaction Number'].str.contains("Adjustments Receipts", na=False)) | 
                        (self.df_assets['Transaction Number'].str.contains("Adjustment: budget increase", na=False)), 'Balance Type'] = 'Income'
        
        # Convert income transactions to negative amounts
        self.df_assets.loc[self.df_assets['Balance Type'] == 'Income', 'Transaction Amount'] = self.df_assets.loc[self.df_assets['Balance Type'] == 'Income', 'Transaction Amount'] * -1

        # Filter out budget transactions and drop columns with None values
        self.df_assets = self.df_assets.loc[(self.df_assets['Balance Type'] != 'Budget')]
        self.df_assets = self.df_assets.drop(columns=[col for col in self.df_assets.columns if col is None])
        self.df_bca = pd.concat([self.df_bca, self.df_assets])

        # Add balances flag
        self.balances = True


    # 30% Completion
    def process_transactions(self) -> None:
        """
        Processes transactions from the BCA dataframe by performing the following:
        
        1. Identifies all commitment transactions with positive amounts.
        2. Adds cluster descriptions for these commitments.
        3. Groups and sorts the commitments using a custom method.
        4. Identifies obligation transactions that have not yet been assigned a cluster.
        5. Groups and sorts these obligations with an offset to maintain unique cluster IDs.

        Notes:
        - Transactions of type 'Budget' are excluded.
        - Only positive 'Commitment' and 'Obligation' amounts are considered.
        - Uses `add_descriptions` to annotate transaction rows.
        - Uses `group_and_sort` to structure grouped outputs with optional offsetting.

        Updates:
        - self.df_bca: updated with descriptions.
        - self.df_sorted: final grouped and sorted transaction data.
        - self.df_reduced: reduced dataset without 'Budget' entries.
        - Prints a message if no commitments or obligations are found.

        Returns:
            None
        """
        # Step 1: Identify commitment transaction numbers with positive amounts
        comm_numbers: np.ndarray = self.df_bca.loc[
            (self.df_bca['Transaction Amount'] > 0) & 
            (self.df_bca['Balance Type'] == 'Commitment'),
            'Transaction Number'
        ].unique()

        # Step 2: Add descriptions for identified commitments
        self.df_bca = self.add_descriptions(self.df_bca, comm_numbers)

        # Step 3: Filter out remaining 'Budget' entries
        self.df_reduced = self.df_bca.loc[self.df_bca['Balance Type'] != 'Budget']

        # Step 4: Group and sort commitments if found
        if comm_numbers.size == 0:
            print("**_Note:_** No new commitments found.")
            self.df_sorted = self.df_reduced
        else:
            self.df_sorted = self.group_and_sort(self.df_reduced, comm_numbers)

        # Step 5: Identify obligation transactions with no cluster assigned
        ob_numbers: np.ndarray = self.df_sorted.loc[
            (self.df_sorted['Cluster'] == '') &
            (self.df_sorted['Balance Type'] == 'Obligation') &
            (self.df_sorted['Transaction Amount'] > 0),
            'Transaction Number'
        ].unique()

        # Step 6: Group and sort obligations if found
        if ob_numbers.size == 0:
            if comm_numbers.size == 0:
                print("**_Note:_** No positive obligations found.")
        else:
            self.df_sorted = self.group_and_sort(self.df_sorted, ob_numbers, offset=len(comm_numbers))


    # 40% Completion
    def process_po(self) -> None:
        """
        Processes Purchase Order (PO) details by performing the following steps:

        1. Adds 'PO Details' to the list of sheet names and the PO file path to tracked file paths.
        2. Extracts specific relevant columns from the original PO DataFrame.
        3. Concatenates rows by PO number using a custom method to combine multiple entries.
        4. Updates `self.df_sorted` with the PO metadata (Item Description, Item Category, Requester Name, Supplier Name)
        for matching PO numbers based on 'Obligation Nr' or 'Transaction Number'.

        Attributes Updated:
        - self.df_po: Filtered PO DataFrame with selected columns.
        - self.df_po_concatenated: Aggregated PO data grouped by PO number.
        - self.df_sorted: Final dataset with enriched PO information for each matching transaction.
        - self.sheet_names: Appends 'PO Details'.
        - self.file_paths: Appends `self.po_details_path`.

        Notes:
        - `concatenate_rows_by_po_number` is expected to group and concatenate rows with the same PO number.
        - Column indices used: [1, 2, 6, 15, 16] must be consistent with the expected PO file structure.

        Returns:
            None
        """
        # Track the sheet and path used
        self.sheet_names.append('PO Details')
        self.file_paths.append(self.po_details_path)

        # Select relevant columns by index
        columns = self.df_po.iloc[:, [1, 2, 6, 15, 16]].columns
        self.df_po = self.df_po[[columns[0], columns[1], columns[2], columns[3], columns[4]]]

        # Group and concatenate PO details by PO number
        self.df_po_concatenated = self.concatenate_rows_by_po_number(self.df_po, group_by=columns[0])

        # Get unique PO numbers
        po_numbers: np.ndarray = self.df_po_concatenated[columns[0]].unique()

        # Update self.df_sorted with PO details based on matching PO number
        for po_number in po_numbers:
            condition = (self.df_sorted['Obligation Nr'] == po_number) | \
                        (self.df_sorted['Transaction Number'] == po_number)

            self.df_sorted.loc[condition, 'Item Description'] = self.df_po_concatenated.loc[
                self.df_po_concatenated[columns[0]] == po_number, columns[3]
            ].values[0]

            self.df_sorted.loc[condition, 'Item Category Description'] = self.df_po_concatenated.loc[
                self.df_po_concatenated[columns[0]] == po_number, columns[4]
            ].values[0]

            self.df_sorted.loc[condition, 'Requester Name'] = self.df_po_concatenated.loc[
                self.df_po_concatenated[columns[0]] == po_number, columns[1]
            ].values[0]

            self.df_sorted.loc[condition, 'Supplier Name'] = self.df_po_concatenated.loc[
                self.df_po_concatenated[columns[0]] == po_number, columns[2]
            ].values[0]

        # Display processing progress
        progress_placeholder.markdown(f"Processing: {40}% complete...")


    def process_accounts(self) -> None:
        """
        Enriches `self.df_sorted` with account descriptions based on transaction account numbers.

        Steps:
        1. Loads a reference Excel file ('Account.numbers.table.xlsx') containing account numbers and their descriptions.
        2. Iterates over each unique 'Transaction Account' in `self.df_sorted`.
        3. Attempts to extract the numeric account identifier from the transaction account string (assumed format: 'X-X-<number>').
        4. Matches this number with the 'Acc No' column from the reference Excel file to get the corresponding description.
        5. If a match is found, populates a temporary column ('Temp') with the description.
        6. Appends the description to both 'Item Description' and 'Item Category Description' columns.
        7. Cleans up the temporary column after enrichment.

        Notes:
        - Assumes 'Transaction Account' has a format that can be split with `-`, with the account number in the third part.
        - If a description lookup fails, it is silently ignored via `try-except`.

        Modifies:
            - self.df_sorted: adds enriched item and category descriptions.

        Returns:
            None
        """
        acc_num_desc: pd.DataFrame = pd.read_excel('Account.numbers.table.xlsx')
        unique_accounts: np.ndarray = self.df_sorted['Transaction Account'].unique()

        for trans_account in unique_accounts:
            if trans_account is not None:
                try:
                    account_number = int(trans_account.split('-')[2])
                    description = acc_num_desc.loc[acc_num_desc['Acc No'] == account_number, 'Account description'].item()
                    self.df_sorted.loc[self.df_sorted['Transaction Account'] == trans_account, 'Temp'] = description
                except Exception:
                    pass  # Skip accounts that fail to parse or look up

        if 'Temp' in self.df_sorted.columns:
            self.df_sorted['Item Description'] = (
                self.df_sorted['Temp'] + ' | ' + self.df_sorted['Item Description'].astype(str)
            )
            self.df_sorted['Item Category Description'] = (
                self.df_sorted['Temp'] + ' | ' + self.df_sorted['Item Category Description'].astype(str)
            )
            self.df_sorted = self.df_sorted.drop(columns=['Temp'])


    # 60% Completion to 100% Completion
    def process_output(self) -> BytesIO:
        """
        Finalizes the transaction data for export by:
        1. Ensuring a consistent column order in `self.df_sorted`.
        2. Adding any missing expected columns with `None` values.
        3. Reordering the dataframe columns according to a predefined `column_order`.
        4. Categorizing and sorting transactions by 'Balance Type' using a logical order.
        5. Setting `self.CFB_flag` based on whether certain opening balance types are present.
        6. Creating and exporting the processed output file using `self.create_output_file`.

        Progress updates are displayed via `progress_placeholder` at 60% and 100% completion.

        Returns:
            The output file as an o
        """    
        column_order = ['Budget Account', 'Cost Center Segment Description', 'Account Description', 
                                   'Transaction Type', 'Transaction SubType', 'Transaction Action', 'Transaction Number', 
                                   'Expense Report Owner', 'Transaction Account', 'Transaction ID', 'Transaction Currency', 
                                   'Activity Type', 'Reservation Amount', 'Liquidation Transaction Type', 'Liquidation Transaction Number', 
                                   'Liquidation Amount', 'Commitment Nr', 'Obligation Nr', 'Expenditure Nr', 
                                   'Cluster', 'Project Code', 'Budget Date', 
                                   'Balance Type', 'Transaction Amount', 'Item Description', 
                                   'Requester Name', 'Supplier Name', 'Item Category Description']
        
        # Ensure all columns are in df_sorted

        for col in column_order:
            if col not in self.df_sorted.columns:
                self.df_sorted[col] = None
        
        existing_columns = [col for col in column_order if col in self.df_sorted.columns]
        
        self.df_sorted = self.df_sorted.reindex(columns=existing_columns)

        progress_placeholder.markdown(f"Current processing: {60}% complete...")

        # Order the transactions to make sense
        categories = []



        if "Opening Balance" in self.df_sorted['Balance Type'].unique():
            categories.append("Opening Balance")
            self.CFB_flag = 1
            
        if "Opening Balance Assets" in self.df_sorted['Balance Type'].unique():
            categories.append("Opening Balance Assets")
            self.CFB_flag = 2

        if "Income" in self.df_sorted['Balance Type'].unique():
            categories.append("Income")
        
        categories.append("Commitment")
        categories.append("Obligation")
        categories.append("Expenditure")

        balance_type_order = pd.CategoricalDtype(
                categories=categories,
                ordered=True
            )

        self.df_sorted["Balance Type"] = self.df_sorted["Balance Type"].astype(balance_type_order)

        self.df_sorted = self.df_sorted.sort_values(by=["Balance Type"], ascending=[True])

        print(self.df_sorted.iloc[0])
        output_file = self.create_output_file(self.df_sorted, self.file_paths)
        progress_placeholder.markdown(f"Current processing: {100}% complete...")

        return output_file


    def auto_process(self) -> str | None:
        """
        Automatically orchestrates the data processing pipeline for BCA, assets, 
        purchase orders, transactions, and account processing. Updates progress 
        at key stages.

        Returns:
            str | None: The path to the output file if processing succeeds, 
                        or None if BCA data is unavailable or validation fails.
        """

        bca_, asts_, po_ = self.extract_and_check_data()

        if bca_:
            self.process_bca()

            progress_placeholder.markdown(f"Current processing: {10}% complete...")

            if asts_:
                self.process_assets()

            progress_placeholder.markdown(f"Current processing: {20}% complete...")

            self.process_transactions()

            progress_placeholder.markdown(f"Current processing: {30}% complete...")

            if po_:
                self.process_po()

            progress_placeholder.markdown(f"Current processing: {40}% complete...")

            self.process_accounts()

            progress_placeholder.markdown(f"Current processing: {50}% complete...")

            output_file = self.process_output()

            return output_file

        else:
            return None

# Streamlit App

## Title

st.title('Making Sense of SUNFIN')

st.markdown('---')

st.markdown('''Version: 2.0''')

st.markdown('''Use the app at your own risk, and please don’t blame us if it does not work or gives the wrong information. 
            You are welcome to improve it by accessing the source code here: [Github](https://github.com/Divanvdb/SUNFIN_app)
''')

## Download Guide to SUNFIN
with open('Guide_to_Making_Sense_of_SunFin.pdf', 'rb') as file:
    pdf_data = file.read()

st.download_button(
    label="Download User Guide",
    data=pdf_data,
    file_name='Guide_to_Making_Sense_of_SunFin.pdf',
    mime='application/pdf'
)

st.markdown('---')

## Sidebar with file upload options

progress_placeholder = st.empty()

st.sidebar.header('Upload Files from Folder')

bca_file, assets_file, po_file = None, None, None

uploaded_files = st.sidebar.file_uploader("Upload Files from Folder", type=["xlsx"], accept_multiple_files=True)

st.sidebar.header('Upload Files Individually')

bca_file_individual = st.sidebar.file_uploader("Upload BCA File", type=["xlsx"])
assets_file_individual = st.sidebar.file_uploader("Upload Assets File", type=["xlsx"])
po_file_individual = st.sidebar.file_uploader("Upload PO Details File", type=["xlsx"])

unique_ids = []
if uploaded_files:
    
    for uploaded_file in uploaded_files:
        extracted_value = uploaded_file.name.split(' - ')[0]
        
        if extracted_value not in unique_ids:
            if '.xlsx' not in extracted_value:
                unique_ids.append(extracted_value)
else:
    bca_file = bca_file_individual
    assets_file = assets_file_individual
    po_file = po_file_individual

## Processing function

if st.sidebar.button('Process'):

    if (bca_file is not None) | (len(unique_ids) >= 1):

        if len(unique_ids) >= 1:
            output_files = []
            output_names = []
            st.write("Multiple files uploaded successfully. Processing will start...")

            st.write(unique_ids)
            for i, unique_id in enumerate(unique_ids):
                bca_file, assets_file, po_file = None, None, None
                for uploaded_file in uploaded_files:
                    if f"{unique_id} -" in uploaded_file.name:
                        if "Assets" in uploaded_file.name or "asset" in uploaded_file.name or 'Assets' in uploaded_file.name:
                            assets_file = uploaded_file
                        if any(f"{i} - BudgetaryControlAnalysis" in uploaded_file.name for i in range(10)) or "BCA" in uploaded_file.name:
                            bca_file = uploaded_file
                        elif "PO" in uploaded_file.name or "PODetails" in uploaded_file.name:
                            po_file = uploaded_file

                st.write(f"**{i + 1} / {len(unique_ids)}** - Files uploaded successfully for **{unique_id}**:")
                if bca_file is not None:
                    st.write(f"- BCA: {bca_file.name}")
                if assets_file is not None:
                    st.write(f"- BCA Assets: {assets_file.name}")
                if po_file is not None:
                    st.write(f"- PO File: {po_file.name}")

                if bca_file:
                    processor = ExcelProcessor(bca_file, assets_file, po_file)
                    progress_placeholder = st.empty()

                    try:
                        output = processor.auto_process()
                        if output is not None:
                            output_files.append(output)
                            
                            output_name = processor.create_file_name(bca_file.name)
                            output_names.append(output_name)

                            st.success(f"**Completed processing for** {unique_id}")
                        else:
                            st.warning(f"**No output for** {unique_id}")
                    except:
                        st.error(f"**Error processing** {unique_id}")
                else:
                    st.error(f"**No BCA file found for** {unique_id}")

            if len(output_files) == 0:
                st.error("No output files were generated.")

            else:
                zip_buffer = BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
                    for output_file, output_name in zip(output_files, output_names):

                        if isinstance(output_file, BytesIO):
                            # Create a temporary file
                            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
                                temp_file.write(output_file.getvalue())  # Write the BytesIO content to the temp file
                                temp_file_path = temp_file.name  # Store the temporary file path

                            # Add the temporary file to the ZIP
                            zip_file.write(temp_file_path, arcname=output_name)

                            # Clean up: Delete the temporary file after adding to the ZIP
                            os.remove(temp_file_path)
                        else:
                            # If output_file is already a file path, just add it directly
                            zip_file.write(output_file, arcname=output_name)

                zip_buffer.seek(0)

                st.download_button(
                    label="Download All Updated Excel Files",
                    data=zip_buffer,
                    file_name="output_files.zip",
                    mime="application/zip"
                )

        else:
            if bca_file is not None:
                st.write(f"Files uploaded successfully:")
                st.write(f"- BCA: {bca_file.name}")

                if assets_file is not None:
                    st.write(f"- BCA Assets: {assets_file.name}")
                if po_file is not None:
                    st.write(f"- PO File: {po_file.name}")

                processor = ExcelProcessor(bca_file, assets_file, po_file)

                output_file = processor.auto_process()

                if output_file is not None:

                    st.write("**Completed processing.**")

                    st.download_button(
                                label="Download Updated Excel",
                                data=output_file,
                                file_name=processor.create_file_name(bca_file.name),
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                else:
                    st.error("No output file was generated.")
            else:
                st.error("Please upload the BCA file.")

    else:
        st.error("Please upload all required files.")
